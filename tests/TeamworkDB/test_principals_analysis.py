import os
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

import openpyxl
import pytest

from cornerstone_automation.utils.excel_utils import read_excel, get_excel_headers


def _clean(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _normalize_header_name(h: str) -> str:
    """Normalize header for robust comparisons (lowercase, remove non-alnum)."""
    if h is None:
        return ""
    return "".join(ch for ch in h.lower() if ch.isalnum())


def _choose_key_header(headers: List[str]) -> str:
    """
    Choose a sensible key header from the list of headers.
    Preference order: common id/name candidates, otherwise the first header.
    """
    candidates = ["id", "principal", "name", "employee", "employeeid", "employee_id"]
    normalized = {h: "".join(ch for ch in (h or "").lower() if ch.isalnum()) for h in headers}
    for cand in candidates:
        for h, nh in normalized.items():
            if cand == nh:
                return h
    # fallback to first header
    return headers[0] if headers else ""


def _build_map_by_key(rows: List[Dict[str, Any]], key_header: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Build a map: key_value -> list of row dicts (handles duplicate keys by storing list).
    """
    m: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        key = _clean(r.get(key_header)) if isinstance(r, dict) else ""
        if key == "":
            # use a special empty-key bucket with incremental suffix to avoid collapse
            key = f"__EMPTY_KEY__::{len(m)}"
        m.setdefault(key, []).append(r)
    return m


def _compare_rows_by_headers(
    r1: Dict[str, Any],
    r2: Dict[str, Any],
    headers: List[str]
) -> List[Tuple[str, Any, Any]]:
    """
    Compare two row dicts across headers. Return list of (header, value1, value2) for differing headers.

    NOTE: Skip validation for the 'report time' column (and any header that normalizes to 'reporttime').
    All other columns are compared.
    """
    diffs = []
    for h in headers:
        # Skip 'report time' column by normalized header name
        if _normalize_header_name(h) == "reporttime":
            continue

        v1 = _clean(r1.get(h) if isinstance(r1, dict) else None)
        v2 = _clean(r2.get(h) if isinstance(r2, dict) else None)
        if v1 != v2:
            diffs.append((h, v1, v2))
    return diffs


def _find_officer_header(headers: List[str]) -> Optional[str]:
    """
    Return header that represents 'Officer Name'.

    Matching strategy:
      1. Prefer exact whole-phrase matches like "Officer Name" (case-insensitive),
         using word boundaries so the phrase is treated as whole words.
      2. Fallback: match headers where normalized text contains both 'officer' and 'name'
         (covers variants like 'OfficerName', "Officer's Name", etc.).
    Returns None if not found.
    """
    import re

    if not headers:
        return None

    # 1) Whole-phrase match: "officer name" as whole words (case-insensitive)
    phrase_re = re.compile(r'\bofficer\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            continue
        if phrase_re.search(str(h)):
            return h

    # 2) Fallback: normalized-token check (existing behavior)
    for h in headers:
        if h is None:
            continue
        nh = _normalize_header_name(h)
        if "officer" in nh and "name" in nh:
            return h

    return None


def test_compare_principals_analysis_workbooks():
    """
    Compare the two workbooks sheet-by-sheet. For each sheet that exists in both workbooks:
      - choose a key header (Officer Name preferred)
      - build key->row maps for each workbook
      - report keys missing in the other workbook
      - for keys present in both, compare values by column and report per-column differences

    Prints detailed information; the test is informational and does not fail by default.
    """
    base = Path(__file__).parent
    fixtures_dir = base / "fixtures"

    file1 = fixtures_dir / "Principal TWDB analysis - fix 20260130.xlsx"
    file2 = fixtures_dir / "Principal TWDB analysis.xlsx"

    assert file1.exists(), f"Missing fixture: {file1}"
    assert file2.exists(), f"Missing fixture: {file2}"

    wb1 = openpyxl.load_workbook(str(file1), data_only=True)
    wb2 = openpyxl.load_workbook(str(file2), data_only=True)

    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)

    only_in_file1 = sorted(sheets1 - sheets2)
    only_in_file2 = sorted(sheets2 - sheets1)
    common_sheets = sorted(sheets1 & sheets2)

    print("\n--- Workbook sheet summary ---")
    print(f"Sheets only in {file1.name}: {only_in_file1 or '(none)'}")
    print(f"Sheets only in {file2.name}: {only_in_file2 or '(none)'}")
    print(f"Sheets in both files: {common_sheets or '(none)'}\n")

    overall_missing = 0
    overall_diffs = 0

    for sheet in common_sheets:
        print(f"\n=== Comparing sheet: '{sheet}' ===")
        headers1 = get_excel_headers(str(file1), sheet_name=sheet)
        headers2 = get_excel_headers(str(file2), sheet_name=sheet)

        # Choose header set to compare (use headers1 if available, otherwise headers2)
        headers = headers1 if headers1 else headers2
        if not headers:
            print(f"  Warning: sheet '{sheet}' has no headers in either file. Skipping.")
            continue

        # Prefer 'Officer Name' as key column if present in either sheet
        officer_header = _find_officer_header(headers1) if headers1 else None
        if officer_header is None:
            officer_header = _find_officer_header(headers2) if headers2 else None

        if officer_header:
            key_header = officer_header
            print(f"  officer header: '{key_header}'")
        else:
            # fallback to previous logic
            print(f"  other header:")
            key_header = _choose_key_header(headers1) if headers1 else _choose_key_header(headers2)
            if not key_header:
                key_header = headers[2]

        print(f"  Using key column: '{key_header}'")
        rows1 = read_excel(str(file1), sheet_name=sheet)
        rows2 = read_excel(str(file2), sheet_name=sheet)

        map1 = _build_map_by_key(rows1, key_header)
        map2 = _build_map_by_key(rows2, key_header)

        keys1 = set(map1.keys())
        keys2 = set(map2.keys())

        missing_in_2 = sorted(k for k in keys1 - keys2)
        missing_in_1 = sorted(k for k in keys2 - keys1)

        if missing_in_2:
            overall_missing += len(missing_in_2)
            print(f"  Rows present in {file1.name} but missing in {file2.name}: {len(missing_in_2)}")
            for k in missing_in_2:
                print(f"    - {k}")
        else:
            print(f"  No rows missing in {file2.name}")

        if missing_in_1:
            overall_missing += len(missing_in_1)
            print(f"  Rows present in {file2.name} but missing in {file1.name}: {len(missing_in_1)}")
            for k in missing_in_1:
                print(f"    - {k}")
        else:
            print(f"  No rows missing in {file1.name}")

        # Compare rows with same key. handle duplicates: compare each combination of rows for that key.
        common_keys = sorted(keys1 & keys2)
        sheet_diffs = 0
        for k in common_keys:
            list1 = map1.get(k, [])
            list2 = map2.get(k, [])
            # Compare pairwise by index; if counts differ, still compare up to min length and report extras as missing
            min_len = min(len(list1), len(list2))
            for i in range(min_len):
                diffs = _compare_rows_by_headers(list1[i], list2[i], headers)
                if diffs:
                    sheet_diffs += 1
                    overall_diffs += 1
                    print(f"\n  Differences for key '{k}' (row #{i+1} for that key):")
                    for col, v1, v2 in diffs:
                        print(f"    Column '{col}': {file1.name}='{v1}'  |  {file2.name}='{v2}'")
            # extra rows for key in one file
            if len(list1) > min_len:
                sheet_diffs += len(list1) - min_len
                overall_missing += len(list1) - min_len
                print(f"\n  Extra {len(list1)-min_len} row(s) for key '{k}' in {file1.name} not present in {file2.name}")
            if len(list2) > min_len:
                sheet_diffs += len(list2) - min_len
                overall_missing += len(list2) - min_len
                print(f"\n  Extra {len(list2)-min_len} row(s) for key '{k}' in {file2.name} not present in {file1.name}")

        if sheet_diffs == 0:
            print(f"  No differing rows found in sheet '{sheet}'")
        else:
            print(f"  Found {sheet_diffs} differing row(s) in sheet '{sheet}'")

    print("\n--- Comparison complete ---")
    print(f"Total missing-row count: {overall_missing}")
    print(f"Total differing-row count: {overall_diffs}")

    # Informational test: do not fail the run by default
    assert True
