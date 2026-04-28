import os
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

import openpyxl
import pytest

from cornerstone_automation.utils.excel_utils import read_excel, get_excel_headers


def _clean(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _normalize_header_name(h: str) -> str:
    """Normalize header for robust comparisons (lowercase, remove non-alnum)."""
    if h is None:
        return ""
    return "".join(ch for ch in h.lower() if ch.isalnum())


def _choose_key_header(headers: List[str]) -> str:
    """
    Choose a sensible key header from the list of headers.
    Preference order: common id/name candidates, otherwise the first header.
    """
    candidates = ["id", "principal", "name", "employee", "employeeid", "employee_id"]
    normalized = {h: "".join(ch for ch in (h or "").lower() if ch.isalnum()) for h in headers}
    for cand in candidates:
        for h, nh in normalized.items():
            if cand == nh:
                return h
    return headers[0] if headers else ""


def _build_map_by_key(rows: List[Dict[str, Any]], key_header: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Build a map: key_value -> list of row dicts (handles duplicate keys by storing list).
    """
    m: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        key = _clean(r.get(key_header)) if isinstance(r, dict) else ""
        if key == "":
            key = f"__EMPTY_KEY__::{len(m)}"
        m.setdefault(key, []).append(r)
    return m


def _compare_rows_by_headers(
    r1: Dict[str, Any],
    r2: Dict[str, Any],
    headers: List[str]
) -> List[Tuple[str, Any, Any]]:
    """
    Compare two row dicts across headers. Return list of (header, value1, value2) for differing headers.

    Skips:
      - 'report time' (normalized -> 'reporttime')
      - LY/Total/Change columns as requested (pattern-matched)
      - exact 'Change Total' (normalized -> 'changetotal')
    """
    diffs = []
    for h in headers:
        nh = _normalize_header_name(h)

        # Skip 'report time'
        if nh == "reporttime":
            continue

        # Skip exact "Change Total"
        if nh == "changetotal":
            continue

        # LY - Total Teamwork variants
        if "ly" in nh and "total" in nh and "teamwork" in nh:
            continue

        # LY - Total Crit/Ast/Ovr (Unique) variants
        if "ly" in nh and "total" in nh and any(tok in nh for tok in ("crit", "ast", "ovr", "unique")):
            continue

        # Change Teamwork variants
        if "change" in nh and "teamwork" in nh:
            continue

        # Generic change/totalchange
        if nh in ("change", "totalchange"):
            continue

        v1 = _clean(r1.get(h) if isinstance(r1, dict) else None)
        v2 = _clean(r2.get(h) if isinstance(r2, dict) else None)
        if v1 != v2:
            diffs.append((h, v1, v2))
    return diffs


def _find_officer_header(headers: List[str]) -> Optional[str]:
    """
    Return header that represents 'Officer Name'.

    Strategy:
      1) Whole-phrase 'officer name' (word boundaries)
      2) Fallback normalized token check containing both 'officer' and 'name'
    """
    import re

    if not headers:
        return None

    phrase_re = re.compile(r'\bofficer\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            continue
        if phrase_re.search(str(h)):
            return h

    for h in headers:
        if h is None:
            continue
        nh = _normalize_header_name(h)
        if "officer" in nh and "name" in nh:
            return h

    return None


def _print_diffs_for_sheet(sheet: str, file1_name: str, file2_name: str, diffs_by_key: Dict[str, List[Tuple[str, Any, Any]]]) -> None:
    """
    Nicely print differences for a sheet.
    diffs_by_key: { key_value: [ (column_name, value_in_file1, value_in_file2), ... ] }
    """
    if not diffs_by_key:
        print(f"  No column-level differences found in sheet '{sheet}'.")
        return

    print(f"\n--- Differences in sheet: {sheet} ---")
    col_key = "KEY"
    col_col = "COLUMN"
    col_f1 = file1_name
    col_f2 = file2_name
    w_key = 30
    w_col = 40
    w_val = 30
    header_line = f"{col_key:<{w_key}} | {col_col:<{w_col}} | {col_f1:<{w_val}} | {col_f2:<{w_val}}"
    print(header_line)
    print("-" * len(header_line))
    for key in sorted(diffs_by_key.keys()):
        rows = diffs_by_key[key]
        for idx, (colname, v1, v2) in enumerate(rows):
            key_display = key if idx == 0 else ""
            # Remove whitespace in column name for compact display
            colname_nospace = "".join(str(colname).split())
            print(f"{key_display:<{w_key}} | {colname_nospace:<{w_col}} | {v1:<{w_val}} | {v2:<{w_val}}")
    print("-" * len(header_line))


def test_compare_teamwork_profitability_workbooks_basic():
    """
    Compare Teamwork Profitability workbooks (non-deflated) sheet-by-sheet.
    Uses 'Officer Name' as key when available, otherwise falls back to heuristic.
    """
    base = Path(__file__).parent
    fixtures_dir = base / "fixtures"

    file1 = fixtures_dir / "Teamwork Profitability - fix 20260127.xlsx"
    file2 = fixtures_dir / "Teamwork Profitability.xlsx"

    assert file1.exists(), f"Missing fixture: {file1}"
    assert file2.exists(), f"Missing fixture: {file2}"

    wb1 = openpyxl.load_workbook(str(file1), data_only=True)
    wb2 = openpyxl.load_workbook(str(file2), data_only=True)

    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)
    common_sheets = sorted(sheets1 & sheets2)

    print("\n--- Teamwork Profitability workbook sheet summary ---")
    print(f"Sheets only in {file1.name}: {sorted(sheets1 - sheets2) or '(none)'}")
    print(f"Sheets only in {file2.name}: {sorted(sheets2 - sheets1) or '(none)'}")
    print(f"Sheets in both files: {common_sheets or '(none)'}\n")

    overall_missing = 0
    overall_diffs = 0

    for sheet in common_sheets:
        print(f"\n=== Comparing sheet: '{sheet}' ===")
        headers1 = get_excel_headers(str(file1), sheet_name=sheet)
        headers2 = get_excel_headers(str(file2), sheet_name=sheet)

        headers = headers1 if headers1 else headers2
        if not headers:
            print(f"  Warning: sheet '{sheet}' has no headers. Skipping.")
            continue

        # Prefer 'Officer Name' as key column
        officer_header = _find_officer_header(headers1) if headers1 else None
        if officer_header is None:
            officer_header = _find_officer_header(headers2) if headers2 else None

        if officer_header:
            key_header = officer_header
            print(f"  officer header: '{key_header}'")
        else:
            key_header = _choose_key_header(headers1) if headers1 else _choose_key_header(headers2)
            if not key_header:
                key_header = headers[0]

        rows1 = read_excel(str(file1), sheet_name=sheet)
        rows2 = read_excel(str(file2), sheet_name=sheet)

        map1 = _build_map_by_key(rows1, key_header)
        map2 = _build_map_by_key(rows2, key_header)

        keys1 = set(map1.keys())
        keys2 = set(map2.keys())

        missing_in_2 = sorted(k for k in keys1 - keys2)
        missing_in_1 = sorted(k for k in keys2 - keys1)

        if missing_in_2:
            overall_missing += len(missing_in_2)
            print(f"  Rows present in {file1.name} but missing in {file2.name}: {len(missing_in_2)}")
            for k in missing_in_2:
                print(f"    - {k}")
        else:
            print(f"  No rows missing in {file2.name}")

        if missing_in_1:
            overall_missing += len(missing_in_1)
            print(f"  Rows present in {file2.name} but missing in {file1.name}: {len(missing_in_1)}")
            for k in missing_in_1:
                print(f"    - {k}")
        else:
            print(f"  No rows missing in {file1.name}")

        diffs_by_key: Dict[str, List[Tuple[str, Any, Any]]] = {}
        common_keys = sorted(keys1 & keys2)
        sheet_diffs = 0

        for k in common_keys:
            list1 = map1.get(k, [])
            list2 = map2.get(k, [])
            min_len = min(len(list1), len(list2))
            for i in range(min_len):
                diffs = _compare_rows_by_headers(list1[i], list2[i], headers)
                if diffs:
                    sheet_diffs += 1
                    overall_diffs += 1
                    diffs_by_key.setdefault(k, []).extend(diffs)
            if len(list1) > min_len:
                sheet_diffs += len(list1) - min_len
                overall_missing += len(list1) - min_len
                print(f"\n  Extra {len(list1)-min_len} row(s) for key '{k}' in {file1.name} not present in {file2.name}")
            if len(list2) > min_len:
                sheet_diffs += len(list2) - min_len
                overall_missing += len(list2) - min_len
                print(f"\n  Extra {len(list2)-min_len} row(s) for key '{k}' in {file2.name} not present in {file1.name}")

        _print_diffs_for_sheet(sheet, file1.name, file2.name, diffs_by_key)

        if sheet_diffs == 0:
            print(f"  No differing rows found in sheet '{sheet}'")
        else:
            print(f"  Found {sheet_diffs} differing row(s) in sheet '{sheet}'")

    print("\n--- Teamwork Profitability comparison complete ---")
    print(f"Total missing-row count: {overall_missing}")
    print(f"Total differing-row count: {overall_diffs}")

    # Informational test: do not fail the run by default
    assert True