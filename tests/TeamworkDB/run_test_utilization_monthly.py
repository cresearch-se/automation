#!/usr/bin/env python3
"""
run_utilization_monthly_tests.py

Runs all utilization QA test groups, parses the output files, and generates
Excel and HTML reports saved to tests/TeamworkDB/output/.

Usage (run from Code/ directory):
    py tests/TeamworkDB/run_utilization_monthly_tests.py
"""

import subprocess
import sys
import re
from pathlib import Path
from datetime import date
from typing import Any

# ==========================================
# CONFIGURATION
# ==========================================

TEST_FILE  = "tests/TeamworkDB/test_utilization_monthly.py"
OUTPUT_DIR = Path("tests/TeamworkDB/output")

TEST_GROUPS = [
    {
        "name": "Format Validations",
        "filter": (
            "test_format_validations_by_location or "
            "test_totals_validation_by_location or "
            "test_format_validations_by_employee"
        ),
        "output_file": OUTPUT_DIR / "format-validations.txt"
    },
    {
        "name": "DB Location",
        "filter": (
            "test_db_comparison_us_monthly or "
            "test_db_comparison_europe_monthly or "
            "test_db_comparison_us_ytd or "
            "test_db_comparison_europe_ytd"
        ),
        "output_file": OUTPUT_DIR / "db-location.txt"
    },
    {
        "name": "DB Employee Monthly",
        "filter": "test_db_comparison_employee_monthly",
        "output_file": OUTPUT_DIR / "db-employee-monthly.txt"
    },
    {
        "name": "DB Employee YTD",
        "filter": "test_db_comparison_employee_ytd",
        "output_file": OUTPUT_DIR / "db-employee-ytd.txt"
    }
]


# Column definitions for bucketed report sections
MISSING_COLS      = ["Type", "EmpNo", "Name"]
FORMAT_ERROR_COLS = ["Type", "Detail"]


# ==========================================
# STEP 1 — RUN TESTS
# ==========================================

def run_tests() -> None:
    """Run all test groups and write output to individual .txt files."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for group in TEST_GROUPS:
        print(f"  Running: {group['name']}...")
        cmd = [
            sys.executable, "-m", "pytest", TEST_FILE,
            "-k", group["filter"],
            "-v", "-p", "no:cov", "-s", "--tb=short",
            "--override-ini=addopts="
        ]
        with open(group["output_file"], "w", encoding="utf-8") as f:
            result = subprocess.run(cmd, stdout=f, stderr=subprocess.STDOUT, text=True)
        status = "PASS" if result.returncode == 0 else "FAILURES FOUND"
        print(f"    → {status}")


# ==========================================
# STEP 2 — PARSE OUTPUT FILES
# ==========================================

def parse_output_file(filepath: Path | str) -> dict[str, Any]:
    """
    Parse a pytest output .txt file and return structured results:
    {
        "passed":       [test_name, ...],
        "failed":       [test_name, ...],
        "mismatches":   [{"sheet": str, "columns": [...], "rows": [{col: val}]}, ...],
        "missing":      [{"sheet": str, "rows": [{"Type": str, "EmpNo": str, "Name": str}]}, ...],
        "other_errors": [{"sheet": str, "rows": [{"Type": str, "Detail": str}]}, ...]
    }
    """
    content = Path(filepath).read_text(encoding="utf-8", errors="replace")
    lines   = content.splitlines()

    passed         = []
    failed         = []
    mismatches     = []
    missing_tables = []   # [{"sheet": str, "rows": [{"Type", "EmpNo", "Name"}]}]
    other_errors   = []   # [{"sheet": str, "rows": [{"Type", "Detail"}]}]

    def _missing_sheet_bucket(sheet):
        """Return existing bucket for sheet or create a new one."""
        for b in missing_tables:
            if b["sheet"] == sheet:
                return b
        bucket = {"sheet": sheet, "rows": []}
        missing_tables.append(bucket)
        return bucket

    def _parse_missing_line(line, sheet):
        """Parse '[MISSING IN DB] EmpNo=X | Name=Y' into a structured row."""
        m_type = re.match(r'\[(MISSING IN [^\]]+)\]', line)
        row_type = m_type.group(1) if m_type else "MISSING"
        empno = re.search(r'EmpNo=(\S+)', line)
        name  = re.search(r'Name=(.+?)(\s*\|.*)?$', line)
        _missing_sheet_bucket(sheet)["rows"].append({
            "Type":  row_type,
            "EmpNo": empno.group(1) if empno else "",
            "Name":  name.group(1).strip() if name else "",
        })

    # Extract pass/fail per test (test name and PASSED/FAILED are on separate lines)
    last_test = None
    for line in lines:
        m = re.search(r'::(test_\S+)', line)
        if m:
            last_test = m.group(1)
        elif line.strip() in ("PASSED", "FAILED") and last_test:
            (passed if line.strip() == "PASSED" else failed).append(last_test)
            last_test = None

    # Extract mismatch tables, missing rows, format errors
    i = 0
    current_sheet = ""
    while i < len(lines):
        line = lines[i].strip()

        # Track current test/sheet name from parametrized test lines e.g. [EU_Utilization_Monthly]
        m_sheet = re.search(r'\[([A-Za-z0-9_]+)\]\s*$', lines[i])
        if m_sheet and '::test_' in lines[i]:
            current_sheet = m_sheet.group(1)

        # VALUE MISMATCH table block
        if "UTILIZATION QA" in line and "VALUE MISMATCHES" in line:
            sheet = ""
            if i + 1 < len(lines):
                m = re.search(r'Sheet:\s*(.+?)\s*\|', lines[i + 1])
                if m:
                    sheet = m.group(1).strip()

            # Find header row; also collect missing rows that appear before the table
            header_idx = None
            for j in range(i + 1, len(lines)):
                jline = lines[j].strip()
                if jline in ("PASSED", "FAILED") or "::test_" in lines[j]:
                    break  # reached next test, no table found
                if jline.startswith("[MISSING IN DB]") or jline.startswith("[MISSING IN XLS]"):
                    _parse_missing_line(jline, sheet)
                if "Column" in lines[j] and "DB" in lines[j] and "Diff" in lines[j]:
                    header_idx = j
                    break

            if header_idx is None:
                i += 1
                continue

            col_names  = [c.strip() for c in lines[header_idx].split("|")]
            data_start = header_idx + 2  # skip separator line
            table_rows = []

            for j in range(data_start, len(lines)):
                data_line = lines[j].strip()
                if not data_line or "|" not in data_line:
                    break
                values = [v.strip() for v in data_line.split("|")]
                if len(values) == len(col_names):
                    table_rows.append(dict(zip(col_names, values)))

            if table_rows:
                mismatches.append({
                    "sheet":   sheet,
                    "columns": col_names,
                    "rows":    table_rows
                })

            i = data_start + len(table_rows)
            continue

        # Missing rows (outside UTILIZATION QA block — use current_sheet context)
        if line.startswith("[MISSING IN DB]") or line.startswith("[MISSING IN XLS]"):
            _parse_missing_line(line, current_sheet)

        # Format validation errors
        if any(line.startswith(tag) for tag in [
            "[MISSING OFFICE]", "[MISSING TITLE]", "[BLANK VALUE]",
            "[DUPLICATE", "[WRONG ORDER]", "[ZERO VALUE]", "[MISSING SUBTOTAL",
            "[MISSING GRAND TOTAL"
        ]):
            m_type = re.match(r'\[([^\]]+)\]\s*(.*)', line)
            err_type   = m_type.group(1) if m_type else ""
            err_detail = m_type.group(2).strip() if m_type else line
            sheet_key  = current_sheet or ""
            bucket = next((b for b in other_errors if b["sheet"] == sheet_key), None)
            if bucket is None:
                bucket = {"sheet": sheet_key, "rows": []}
                other_errors.append(bucket)
            bucket["rows"].append({"Type": err_type, "Detail": err_detail})

        i += 1

    return {
        "passed":       passed,
        "failed":       failed,
        "mismatches":   mismatches,
        "missing":      missing_tables,
        "other_errors": other_errors
    }


def parse_all_groups() -> list[dict[str, Any]]:
    """Parse all output files and attach group name to each result."""
    results = []
    for group in TEST_GROUPS:
        if not Path(group["output_file"]).exists():
            results.append({"name": group["name"], "passed": [], "failed": [],
                             "mismatches": [], "missing": [], "other_errors": [],
                             "error": "Output file not found"})
            continue
        parsed       = parse_output_file(group["output_file"])
        parsed["name"] = group["name"]
        results.append(parsed)
    return results


def get_report_name() -> str:
    """Read FIXTURE_FILE from the test file and derive the report name."""
    try:
        with open(TEST_FILE, encoding="utf-8") as f:
            for line in f:
                m = re.search(r'FIXTURE_FILE\s*=\s*["\'].*?([^/\\]+\.xlsx)["\']', line)
                if m:
                    month = m.group(1).replace("Utilization_", "").replace(".xlsx", "")
                    return f"utilization_failures_report_{month}"
    except Exception:
        pass
    return f"utilization_failures_report_{date.today().strftime('%Y%m')}"


# ==========================================
# STEP 3A — EXCEL REPORT
# ==========================================

def generate_excel_report(results: list[dict[str, Any]], report_name: str) -> None:
    """Generate an Excel report with one sheet per test group."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [SKIP] openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font   = Font(bold=True)
    title_font  = Font(bold=True, size=13)
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_header_row(ws, values):
        ws.append(values)
        row = ws.max_row
        for col_idx in range(1, len(values) + 1):
            cell = ws.cell(row, col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)

    def write_bucket_section(ws, title, buckets, cols):
        """Write a titled section of bucketed rows as red-highlighted tables, one per sheet."""
        ws.append([title])
        ws.cell(ws.max_row, 1).font = bold_font
        ws.append([])
        for bucket in buckets:
            ws.append([f"Sheet: {bucket['sheet']}"])
            ws.cell(ws.max_row, 1).font = bold_font
            set_header_row(ws, cols)
            for row_data in bucket["rows"]:
                ws.append([row_data.get(c, "") for c in cols])
                row_num = ws.max_row
                for col_idx in range(1, len(cols) + 1):
                    ws.cell(row_num, col_idx).fill   = red_fill
                    ws.cell(row_num, col_idx).border = border
            ws.append([])

    # ---- One sheet per group ----
    for r in results:
        sheet_name = r["name"][:31]
        ws = wb.create_sheet(sheet_name)
        passed = len(r.get("passed", []))
        failed = len(r.get("failed", []))

        ws.append([r["name"]])
        ws["A1"].font = title_font
        ws.append([f"Run: {date.today().strftime('%Y-%m-%d')}   |   Passed: {passed}   |   Failed: {failed}"])
        ws.append([])

        # Value mismatch tables
        for block in r.get("mismatches", []):
            ws.append([f"Sheet: {block['sheet']}"])
            ws.cell(ws.max_row, 1).font = bold_font
            set_header_row(ws, block["columns"])
            for row_data in block["rows"]:
                ws.append([row_data.get(c, "") for c in block["columns"]])
                row_num = ws.max_row
                for col_idx in range(1, len(block["columns"]) + 1):
                    ws.cell(row_num, col_idx).fill   = red_fill
                    ws.cell(row_num, col_idx).border = border
            ws.append([])

        if r.get("missing"):
            write_bucket_section(ws, "MISSING ROWS", r["missing"], MISSING_COLS)

        if r.get("other_errors"):
            write_bucket_section(ws, "FORMAT ERRORS", r["other_errors"], FORMAT_ERROR_COLS)

        auto_width(ws)

    out_path = OUTPUT_DIR / f"{report_name}.xlsx"
    wb.save(out_path)
    print(f"  Excel report: {out_path}")


# ==========================================
# STEP 3B — HTML REPORT
# ==========================================

def generate_html_report(results: list[dict[str, Any]], report_name: str) -> None:
    """Generate a self-contained, collapsible HTML report."""
    run_date      = date.today().strftime("%Y-%m-%d")
    total_passed  = sum(len(r.get("passed", [])) for r in results)
    total_failed  = sum(len(r.get("failed", [])) for r in results)
    overall       = "PASS" if total_failed == 0 else "FAIL"
    overall_color = "#27ae60" if overall == "PASS" else "#e74c3c"

    def badge(status):
        color = "#27ae60" if status == "PASS" else "#e74c3c"
        return (f'<span style="background:{color};color:white;padding:2px 10px;'
                f'border-radius:4px;font-weight:bold;font-size:0.85em">{status}</span>')

    def bucket_section_html(title, buckets, cols, count_label):
        html = f'<p class="sheet-label"><strong>{title}</strong></p>'
        for bucket in buckets:
            headers = "".join(f"<th>{c}</th>" for c in cols)
            body    = "".join(
                "<tr>" + "".join(f"<td>{row.get(c, '')}</td>" for c in cols) + "</tr>"
                for row in bucket["rows"]
            )
            html += (
                f'<p class="sheet-label">Sheet: <strong>{bucket["sheet"]}</strong>'
                f'&nbsp;({len(bucket["rows"])} {count_label})</p>'
                f"<table><thead><tr>{headers}</tr></thead><tbody>{body}</tbody></table>"
            )
        return html

    def mismatch_table_html(block):
        cols    = block["columns"]
        headers = "".join(f"<th>{c}</th>" for c in cols)
        body    = "".join(
            "<tr>" + "".join(f"<td>{row.get(c, '')}</td>" for c in cols) + "</tr>"
            for row in block["rows"]
        )
        return f"""
            <p class="sheet-label">Sheet: <strong>{block['sheet']}</strong>
            &nbsp;({len(block['rows'])} mismatch(es))</p>
            <table><thead><tr>{headers}</tr></thead><tbody>{body}</tbody></table>"""

    group_sections = ""
    for r in results:
        passed   = len(r.get("passed", []))
        failed   = len(r.get("failed", []))
        status   = "PASS" if failed == 0 else "FAIL"
        bg_color = "#eafaf1" if status == "PASS" else "#fdedec"
        border_c = "#27ae60" if status == "PASS" else "#e74c3c"

        body_html = ""

        for block in r.get("mismatches", []):
            body_html += mismatch_table_html(block)

        if r.get("missing"):
            body_html += bucket_section_html("Missing Rows", r["missing"], MISSING_COLS, "missing")

        if r.get("other_errors"):
            body_html += bucket_section_html("Format Errors", r["other_errors"], FORMAT_ERROR_COLS, "error(s)")

        if not body_html:
            body_html = "<p class='all-pass'>&#10003; All checks passed.</p>"

        group_sections += f"""
        <details open>
          <summary style="background:{bg_color};border-left:5px solid {border_c};
                          padding:10px 16px;cursor:pointer;font-size:1.05em;font-weight:bold">
            {r['name']} &nbsp;&mdash;&nbsp; Passed: {passed} &nbsp;|&nbsp; Failed: {failed}
            &nbsp;&nbsp;{badge(status)}
          </summary>
          <div class="group-body" style="border:1px solid {border_c};border-top:none">
            {body_html}
          </div>
        </details>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Utilization QA — {report_name}</title>
  <style>
    body        {{ font-family: Arial, sans-serif; margin: 30px; color: #333; background: #f9f9f9; }}
    h1          {{ color: #2c3e50; margin-bottom: 5px; }}
    .summary    {{ background: white; border: 1px solid #ddd; border-radius: 6px;
                   padding: 15px 20px; margin-bottom: 25px; font-size: 0.95em; }}
    .summary span {{ margin-right: 20px; }}
    details      {{ margin-bottom: 10px; border-radius: 4px; overflow: hidden; }}
    summary      {{ list-style: none; user-select: none; }}
    summary::-webkit-details-marker {{ display: none; }}
    .group-body  {{ padding: 15px 20px; background: white; }}
    table        {{ border-collapse: collapse; width: 100%; margin-bottom: 15px;
                   font-size: 0.88em; }}
    th           {{ background: #2f5496; color: white; padding: 7px 12px; text-align: left; }}
    td           {{ padding: 6px 12px; border: 1px solid #ddd; }}
    tr:nth-child(even) td {{ background: #fef9f9; }}
    tr:nth-child(odd)  td {{ background: #fff; }}
    .sheet-label {{ margin: 12px 0 4px; color: #555; }}
    .all-pass    {{ color: #27ae60; font-weight: bold; }}
  </style>
</head>
<body>
  <h1>Utilization QA Report</h1>
  <div class="summary">
    <span><strong>Report:</strong> {report_name}</span>
    <span><strong>Run Date:</strong> {run_date}</span>
    <span><strong>Total Passed:</strong> {total_passed}</span>
    <span><strong>Total Failed:</strong> {total_failed}</span>
    <span><strong>Overall:</strong>
      <span style="color:{overall_color};font-weight:bold">{overall}</span>
    </span>
  </div>
  {group_sections}
</body>
</html>"""

    out_path = OUTPUT_DIR / f"{report_name}.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"  HTML  report: {out_path}")


# ==========================================
# MAIN
# ==========================================

if __name__ == "__main__":
    print("=" * 55)
    print("  UTILIZATION QA TEST RUNNER")
    print("=" * 55)

    print("\n[1/3] Running tests...")
    run_tests()

    print("\n[2/3] Parsing output files...")
    results = parse_all_groups()

    report_name = get_report_name()
    print(f"\n[3/3] Generating reports ({report_name})...")
    generate_excel_report(results, report_name)
    generate_html_report(results, report_name)

    total_failed = sum(len(r.get("failed", [])) for r in results)
    print("\n" + "=" * 55)
    if total_failed == 0:
        print("  ALL TESTS PASSED")
    else:
        print(f"  {total_failed} TEST(S) FAILED — check reports in {OUTPUT_DIR}")
    print("=" * 55)
