import math
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pytest

from cornerstone_automation.utils.excel_utils import read_excel, get_excel_headers


def _clean(v: Any) -> str:
    return "" if v is None else str(v).strip()

def _normalize_header_name(h: Optional[str]) -> str:
    if h is None:
        return ""
    return "".join(ch for ch in h.lower() if ch.isalnum())

def _find_key_header(headers: List[str]) -> Optional[str]:
    """
    Find a sensible key header for rows:
      - prefer Officer Name
      - then Employee Name
      - then a column named exactly "Name"
      - then fall back to Employer/Organization variants
    """
    import re

    if not headers:
        return None

    officer_re = re.compile(r'\bofficer\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            continue
        if officer_re.search(str(h)):
            return h

    for h in headers:
        if h is None:
            continue
        nh = _normalize_header_name(h)
        if "officer" in nh and "name" in nh:
            return h

    employee_re = re.compile(r'\bemployee\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            continue
        if employee_re.search(str(h)):
            return h

    for h in headers:
        if h is None:
            continue
        if str(h).strip().lower() == "name":
            return h

    employer_re = re.compile(r'\bemployer\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            continue
        if employer_re.search(str(h)):
            return h

    for h in headers:
        if h is None:
            continue
        nh = _normalize_header_name(h)
        if "employer" in nh or "organization" in nh or ("org" in nh and "name" in nh) or ("organizationname" in nh):
            return h

    return None


def _find_rank_header(headers: List[str]) -> Optional[str]:
    """
    Find a sensible rank-like header:
      - Rank Description variants
      - Employee Job Desc / Job Description variants
    """
    if not headers:
        return None

    import re

    patterns = [
        r'\brank\s+description\b',
        r'\brank\s+desc\b',
        r'\brankdescription\b',
        r'\bemployee\s+job\s+desc\b',
        r'\bemployee\s+job\s+description\b',
        r'\bjob\s+description\b',
        r'\bjob\s+desc\b',
    ]
    for pat in patterns:
        cre = re.compile(pat, re.IGNORECASE)
        for h in headers:
            if h is None:
                continue
            if cre.search(str(h)):
                return h

    for h in headers:
        if h is None:
            continue
        nh = _normalize_header_name(h)
        if nh == "rankdescription" or ("rank" in nh and "description" in nh):
            return h
        if ("employeejob" in nh and ("desc" in nh or "description" in nh)) or (
            "job" in nh and ("desc" in nh or "description" in nh)
        ):
            return h

    return None


def _build_map_by_key(rows: List[Dict[str, Any]], key_header: str) -> Dict[str, List[Dict[str, Any]]]:
    m: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        key = _clean(r.get(key_header)) if isinstance(r, dict) else ""
        if key == "":
            key = f"__EMPTY_KEY__::{len(m)}"
        m.setdefault(key, []).append(r)
    return m


def _is_numeric_str(x: str) -> bool:
    try:
        if x == "" or x is None:
            return False
        float(x)
        return True
    except Exception:
        return False


def _is_zero_equivalent(x: str) -> bool:
    if x is None:
        return True
    s = str(x).strip()
    if s == "":
        return True
    try:
        return math.isclose(float(s), 0.0, rel_tol=1e-12, abs_tol=1e-9)
    except Exception:
        return False


def _compare_rows_by_headers(
    r1: Dict[str, Any],
    r2: Dict[str, Any],
    headers: List[str],
    file1_name: str,
    file2_name: str,
    sheet: str,
    key: str,
    rank_header: Optional[str],
) -> List[str]:
    diffs: List[str] = []

    rank_val = ""
    if rank_header and isinstance(r1, dict):
        rank_val = _clean(r1.get(rank_header))
    if not rank_val and rank_header and isinstance(r2, dict):
        rank_val = _clean(r2.get(rank_header))

    for h in headers:
        if _normalize_header_name(h) == "reporttime":
            continue
        if _normalize_header_name(h) == "changetotal":
            continue
        if "ly" in _normalize_header_name(h) and "total" in _normalize_header_name(h) and "teamwork" in _normalize_header_name(h):
            continue
        if "ly" in _normalize_header_name(h) and "total" in _normalize_header_name(h) and any(tok in _normalize_header_name(h) for tok in ("crit", "ast", "ovr", "unique")):
            continue
        if "change" in _normalize_header_name(h) and "teamwork" in _normalize_header_name(h):
            continue
        if _normalize_header_name(h) in ("change", "totalchange"):
            continue

        v1_raw = r1.get(h) if isinstance(r1, dict) else None
        v2_raw = r2.get(h) if isinstance(r2, dict) else None
        v1 = _clean(v1_raw)
        v2 = _clean(v2_raw)

        if _is_zero_equivalent(v1) and _is_zero_equivalent(v2):
            continue

        if _is_numeric_str(v1) and _is_numeric_str(v2):
            try:
                if math.isclose(float(v1), float(v2), rel_tol=1e-9, abs_tol=1e-6):
                    continue
            except Exception:
                pass

        if (_is_zero_equivalent(v1) and _is_numeric_str(v2) and math.isclose(float(v2), 0.0, rel_tol=1e-9, abs_tol=1e-6)) or (
            _is_zero_equivalent(v2) and _is_numeric_str(v1) and math.isclose(float(v1), 0.0, rel_tol=1e-9, abs_tol=1e-6)
        ):
            continue

        if v1 != v2:
            diffs.append(
                f"{sheet} | Key='{key}' | RankDescription='{rank_val}' | Column='{h}' | {file1_name}='{v1}' | {file2_name}='{v2}'"
            )
    return diffs


def test_twdb_analysis_deflated_vs_baseline():
    """
    Compare two 'Deflated' TWDB Analysis workbooks sheet-by-sheet.
    Prints missing rows and per-column mismatches including Rank / Employee Job Description.
    Informational only (does not fail the test).
    """
    base = Path(__file__).parent
    fixtures_dir = base / "fixtures"

    file_fix = fixtures_dir / "TWDB Analysis-Deflated-fix 20260126.xlsx"
    file_base = fixtures_dir / "TWDB Analysis-Deflated.xlsx"

    assert file_base.exists(), f"Missing baseline fixture: {file_base}"
    assert file_fix.exists(), f"Missing fix fixture: {file_fix}"

    wb_base = openpyxl.load_workbook(str(file_base), data_only=True)
    wb_fix = openpyxl.load_workbook(str(file_fix), data_only=True)

    sheets_base = set(wb_base.sheetnames)
    sheets_fix = set(wb_fix.sheetnames)

    common_sheets = sorted(sheets_base & sheets_fix)
    assert common_sheets, "No common sheets to compare"

    errors: List[str] = []

    for sheet in common_sheets:
        headers_base = get_excel_headers(str(file_base), sheet_name=sheet) or []
        headers_fix = get_excel_headers(str(file_fix), sheet_name=sheet) or []

        headers_union = list(dict.fromkeys((headers_base or []) + (headers_fix or [])))

        key_header = _find_key_header(headers_base) or _find_key_header(headers_fix)
        if not key_header:
            key_header = headers_union[0] if headers_union else None

        rank_header = _find_rank_header(headers_base) or _find_rank_header(headers_fix)

        rows_base = read_excel(str(file_base), sheet_name=sheet)
        rows_fix = read_excel(str(file_fix), sheet_name=sheet)

        map_base = _build_map_by_key(rows_base, key_header) if key_header else {}
        map_fix = _build_map_by_key(rows_fix, key_header) if key_header else {}

        keys_base = set(map_base.keys())
        keys_fix = set(map_fix.keys())

        missing_in_fix = sorted(keys_base - keys_fix)
        missing_in_base = sorted(keys_fix - keys_base)

        for k in missing_in_fix:
            rank_val = ""
            rows_for_k = map_base.get(k, [])
            if rank_header and rows_for_k:
                rank_val = _clean(rows_for_k[0].get(rank_header))
            errors.append(f"{sheet} | MISSING in Fix | Key='{k}' | RankDescription='{rank_val}'")

        for k in missing_in_base:
            rank_val = ""
            rows_for_k = map_fix.get(k, [])
            if rank_header and rows_for_k:
                rank_val = _clean(rows_for_k[0].get(rank_header))
            errors.append(f"{sheet} | MISSING in Baseline | Key='{k}' | RankDescription='{rank_val}'")

        common_keys = sorted(keys_base & keys_fix)
        for k in common_keys:
            list_base = map_base.get(k, [])
            list_fix = map_fix.get(k, [])
            min_len = min(len(list_base), len(list_fix))
            for i in range(min_len):
                diffs = _compare_rows_by_headers(
                    list_base[i],
                    list_fix[i],
                    headers_union,
                    file_base.name,
                    file_fix.name,
                    sheet,
                    k,
                    rank_header,
                )
                errors.extend(diffs)
            if len(list_base) > min_len:
                for extra in list_base[min_len:]:
                    rv = _clean(extra.get(rank_header)) if rank_header else ""
                    errors.append(f"{sheet} | EXTRA row in Baseline for Key='{k}' | RankDescription='{rv}'")
            if len(list_fix) > min_len:
                for extra in list_fix[min_len:]:
                    rv = _clean(extra.get(rank_header)) if rank_header else ""
                    errors.append(f"{sheet} | EXTRA row in Fix for Key='{k}' | RankDescription='{rv}'")

    if errors:
        print("\n--- TWDB Analysis (Deflated) comparison - discrepancies found ---")
        print(f"Baseline: {file_base}")
        print(f"Fix:      {file_fix}")
        print(f"Total discrepancies: {len(errors)}\n")
        for e in errors:
            print(e)
    else:
        print("\n--- TWDB Analysis (Deflated) comparison - no discrepancies found ---")

    assert True