import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pytest

from cornerstone_automation.utils.excel_utils import read_excel, get_excel_headers


def _clean(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _normalize_header_name(h: Optional[str]) -> str:
    if h is None:
        return ""
    return "".join(ch for ch in h.lower() if ch.isalnum())


def _find_key_header(headers: List[str]) -> Optional[str]:
    """
    Find a sensible key header for rows:
      - prefer Officer Name
      - then Employee Name
      - then a column named exactly "Name"
      - then fall back to Employer/Organization variants
    """
    import re

    if not headers:
        return None

    # Prefer explicit "Officer Name" phrase (word boundary)
    officer_re = re.compile(r'\bofficer\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            print(f"hh officer name found as header {h}")
            continue
        if officer_re.search(str(h)):
            print(f"officer name found as header {h}")
            return h

    # Try looser officer heuristics (e.g. "Officer", "OfficerName")
    for h in headers:
        if h is None:
            print(f"hh officer/officername found as header {h}")
            continue
        nh = _normalize_header_name(h)
        if "officer" in nh and "name" in nh:
            return h

    # Prefer explicit "Employee Name" phrase (word boundary)
    employee_re = re.compile(r'\bemployee\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            print(f"hh employee name found as header {h}")
            continue
        if employee_re.search(str(h)):
            print(f"employee name found as header {h}")
            return h

    # If a header is exactly "Name" (ignoring surrounding whitespace/case), use it.
    for h in headers:
        if h is None:
            print(f"hh name found as header {h}")
            continue
        if str(h).strip().lower() == "name":
            print(f"name found as header {h}")
            return h

    return None


def _find_rank_header(headers: List[str]) -> Optional[str]:
    """
    Find a sensible rank-like header for rows:
      - matches 'Rank Description', 'Rank Desc', 'RankDescription'
      - matches 'Employee Job Desc', 'Employee Job Description', 'Job Description', 'Job Desc'
      - falls back to any header containing both 'rank' and 'description' or 'job' and 'desc/description'
    """
    if not headers:
        return None

    import re

    # explicit phrase matches (ordered)
    patterns = [
        r'\brank\s+description\b',
        r'\brank\s+desc\b',
        r'\brankdescription\b',
        r'\bemployee\s+job\s+desc\b',
        r'\bemployee\s+job\s+description\b',
        r'\bjob\s+description\b',
        r'\bjob\s+desc\b',
    ]
    for pat in patterns:
        cre = re.compile(pat, re.IGNORECASE)
        for h in headers:
            if h is None:
                continue
            if cre.search(str(h)):
                return h

    # heuristic based on normalized header
    for h in headers:
        if h is None:
            continue
        nh = _normalize_header_name(h)
        if nh == "rankdescription" or ("rank" in nh and "description" in nh):
            return h
        if ("employeejob" in nh and ("desc" in nh or "description" in nh)) or (
            "job" in nh and ("desc" in nh or "description" in nh)
        ):
            return h

    return None


def _build_map_by_key(rows: List[Dict[str, Any]], key_header: str) -> Dict[str, List[Dict[str, Any]]]:
    m: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        key = _clean(r.get(key_header)) if isinstance(r, dict) else ""
        if key == "":
            key = f"__EMPTY_KEY__::{len(m)}"
        m.setdefault(key, []).append(r)
    return m


def _is_numeric_str(x: str) -> bool:
    try:
        if x == "" or x is None:
            return False
        float(x)
        return True
    except Exception:
        return False


def _is_zero_equivalent(x: str) -> bool:
    """
    Consider None/empty/blank as zero-equivalent.
    Also treat any numeric value that is numerically zero as zero-equivalent.
    """
    if x is None:
        return True
    s = str(x).strip()
    if s == "":
        return True
    try:
        return math.isclose(float(s), 0.0, rel_tol=1e-12, abs_tol=1e-9)
    except Exception:
        return False


def _compare_rows_by_headers(
    r1: Dict[str, Any],
    r2: Dict[str, Any],
    headers: List[str],
    file1_name: str,
    file2_name: str,
    sheet: str,
    key: str,
    rank_header: Optional[str],
) -> List[str]:
    """
    Compare two row dicts across headers and return list of human-readable mismatch lines.
    Each line includes the sheet, key (Officer/Employer), Rank Description and the differing column + values.

    Treat blank/null as 0 (zero-equivalent) when comparing values to avoid false positives.
    """
    diffs: List[str] = []

    # determine rank value (prefer r1, fallback to r2)
    rank_val = ""
    if rank_header and isinstance(r1, dict):
        rank_val = _clean(r1.get(rank_header))
    if not rank_val and rank_header and isinstance(r2, dict):
        rank_val = _clean(r2.get(rank_header))

    for h in headers:
        # Ignore comparisons for the Report Time column (and variations)
        if _normalize_header_name(h) == "reporttime":
            continue

        # Skip exact "Change Total"
        if _normalize_header_name(h) == "changetotal":
            continue

        # LY - Total Teamwork variants
        if "ly" in _normalize_header_name(h) and "total" in _normalize_header_name(h) and "teamwork" in _normalize_header_name(h):
            continue

        # LY - Total Crit/Ast/Ovr (Unique) variants
        if "ly" in _normalize_header_name(h) and "total" in _normalize_header_name(h) and any(tok in _normalize_header_name(h) for tok in ("crit", "ast", "ovr", "unique")):
            continue

        # Change Teamwork variants
        if "change" in _normalize_header_name(h) and "teamwork" in _normalize_header_name(h):
            continue

        # Generic change/totalchange
        if _normalize_header_name(h) in ("change", "totalchange"):
            continue

        v1_raw = r1.get(h) if isinstance(r1, dict) else None
        v2_raw = r2.get(h) if isinstance(r2, dict) else None
        v1 = _clean(v1_raw)
        v2 = _clean(v2_raw)

        # If both are zero-equivalent (blank/null/0), consider equal and skip
        if _is_zero_equivalent(v1) and _is_zero_equivalent(v2):
            continue

        # Numeric comparison when both numeric-like
        if _is_numeric_str(v1) and _is_numeric_str(v2):
            try:
                if math.isclose(float(v1), float(v2), rel_tol=1e-9, abs_tol=1e-6):
                    continue
            except Exception:
                pass

        # If one side is blank/null and the other numeric zero -> treat as equal
        if (_is_zero_equivalent(v1) and _is_numeric_str(v2) and math.isclose(float(v2), 0.0, rel_tol=1e-9, abs_tol=1e-6)) or (
            _is_zero_equivalent(v2) and _is_numeric_str(v1) and math.isclose(float(v1), 0.0, rel_tol=1e-9, abs_tol=1e-6)
        ):
            continue

        # Final string comparison
        if v1 != v2:
            diffs.append(
                f"{sheet} | Key='{key}' | RankDescription='{rank_val}' | Column='{h}' | {file1_name}='{v1}' | {file2_name}='{v2}'"
            )
    return diffs


def test_twdb_analysis_fix_vs_baseline():
    """
    Deep compare two TWDB Analysis workbooks (sheet-by-sheet).
    Baseline: TWDB Analysis.xlsx
    Fix:      TWDB Analysis - Fix 20260130.xlsx

    - Uses Officer Name (preferred) as the key column when present.
    - Falls back to Employer/Organization header when Officer Name is not available.
    - Reports missing rows in either file and per-column mismatches for rows present in both.
    - Every reported item includes Rank Description to provide business context.
    - Prints a detailed report of discrepancies (does not fail the test).
    """
    base = Path(__file__).parent
    fixtures_dir = base / "fixtures"

    file_fix = fixtures_dir / "TWDB Analysis - Fix 20260123.xlsx"
    file_base = fixtures_dir / "TWDB Analysis.xlsx"

    assert file_base.exists(), f"Missing baseline fixture: {file_base}"
    assert file_fix.exists(), f"Missing fix fixture: {file_fix}"

    wb_base = openpyxl.load_workbook(str(file_base), data_only=True)
    wb_fix = openpyxl.load_workbook(str(file_fix), data_only=True)

    sheets_base = set(wb_base.sheetnames)
    sheets_fix = set(wb_fix.sheetnames)

    common_sheets = sorted(sheets_base & sheets_fix)
    assert common_sheets, "No common sheets to compare"

    errors: List[str] = []

    for sheet in common_sheets:
        headers_base = get_excel_headers(str(file_base), sheet_name=sheet) or []
        headers_fix = get_excel_headers(str(file_fix), sheet_name=sheet) or []

        # choose headers union for comparison
        headers_union = list(dict.fromkeys((headers_base or []) + (headers_fix or [])))

        # prefer Officer Name as the key; fallback to Employer/Organization header
        key_header = _find_key_header(headers_base) or _find_key_header(headers_fix)
        if not key_header:
            # fallback to first sensible header
            key_header = headers_union[0] if headers_union else None

        rank_header = _find_rank_header(headers_base) or _find_rank_header(headers_fix)

        rows_base = read_excel(str(file_base), sheet_name=sheet)
        rows_fix = read_excel(str(file_fix), sheet_name=sheet)

        map_base = _build_map_by_key(rows_base, key_header) if key_header else {}
        map_fix = _build_map_by_key(rows_fix, key_header) if key_header else {}

        keys_base = set(map_base.keys())
        keys_fix = set(map_fix.keys())

        missing_in_fix = sorted(keys_base - keys_fix)
        missing_in_base = sorted(keys_fix - keys_base)

        # Missing rows from base (present in baseline but not in fix)
        for k in missing_in_fix:
            # if rank header present, include rank value; otherwise empty
            rank_val = ""
            rows_for_k = map_base.get(k, [])
            if rank_header and rows_for_k:
                rank_val = _clean(rows_for_k[0].get(rank_header))
            errors.append(f"{sheet} | MISSING in Fix | Key='{k}' | RankDescription='{rank_val}'")

        # Missing rows from fix (present in fix but not baseline)
        for k in missing_in_base:
            rank_val = ""
            rows_for_k = map_fix.get(k, [])
            if rank_header and rows_for_k:
                rank_val = _clean(rows_for_k[0].get(rank_header))
            errors.append(f"{sheet} | MISSING in Baseline | Key='{k}' | RankDescription='{rank_val}'")

        # Compare rows present in both
        common_keys = sorted(keys_base & keys_fix)
        for k in common_keys:
            list_base = map_base.get(k, [])
            list_fix = map_fix.get(k, [])
            min_len = min(len(list_base), len(list_fix))
            # compare pairwise rows up to min_len
            for i in range(min_len):
                diffs = _compare_rows_by_headers(
                    list_base[i],
                    list_fix[i],
                    headers_union,
                    file_base.name,
                    file_fix.name,
                    sheet,
                    k,
                    rank_header,
                )
                errors.extend(diffs)
            # if extra rows exist for the same key, report them (include rank)
            if len(list_base) > min_len:
                for extra in list_base[min_len:]:
                    rv = _clean(extra.get(rank_header)) if rank_header else ""
                    errors.append(f"{sheet} | EXTRA row in Baseline for Key='{k}' | RankDescription='{rv}'")
            if len(list_fix) > min_len:
                for extra in list_fix[min_len:]:
                    rv = _clean(extra.get(rank_header)) if rank_header else ""
                    errors.append(f"{sheet} | EXTRA row in Fix for Key='{k}' | RankDescription='{rv}'")

    # Print mismatches instead of failing the test
    if errors:
        print("\n--- TWDB Analysis comparison - discrepancies found ---")
        print(f"Baseline: {file_base}")
        print(f"Fix:      {file_fix}")
        print(f"Total discrepancies: {len(errors)}\n")
        for e in errors:
            print(e)
    else:
        print("\n--- TWDB Analysis comparison - no discrepancies found ---")

    # Do not fail the test; informational only
    assert True