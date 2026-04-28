"""Excel utility functions."""

import openpyxl
import math
import re
from typing import List, Dict, Optional, Any


def read_excel(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, str]]:
    """
    Parses the given Excel file and returns data as a list of dictionaries.
    Each dictionary represents a row, with keys as column headers.

    :param file_path: Path to the Excel file (.xlsx)
    :param sheet_name: Optional sheet name to parse. Defaults to the first sheet.
    :return: List of dictionaries containing row data.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = [
        {headers[i]: row[i] for i in range(len(headers))}
        for row in rows[1:]
    ]
    return data


def get_excel_headers(file_path: str, sheet_name: Optional[str] = None) -> List[str]:
    """
    Returns the header row for the given Excel sheet as a list of strings.
    Empty header cells are returned as empty strings.

    :param file_path: Path to the Excel file (.xlsx)
    :param sheet_name: Optional sheet name to read. Defaults to the first sheet.
    :return: List of header names (strings). Returns an empty list if the sheet has no rows.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    return headers


def get_excel_row_count(
    file_path: str,
    sheet_name: Optional[str] = None,
    include_header: bool = False,
    non_empty_only: bool = True
) -> int:
    """
    Returns the number of rows in the given Excel sheet.

    Defaults:
      - include_header=False -> counts only data rows (excludes header).
      - non_empty_only=True -> counts only rows that contain at least one non-empty cell.

    :param file_path: Path to the Excel file (.xlsx)
    :param sheet_name: Optional sheet name to read. Defaults to the first sheet.
    :param include_header: If True, include the header row in the count.
    :param non_empty_only: If True, only count rows that have at least one non-empty cell.
    :return: Number of rows (int).
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    target_rows = rows if include_header else rows[1:]

    if not non_empty_only:
        return len(target_rows)

    count = 0
    for row in target_rows:
        if any(
            cell is not None and (not isinstance(cell, str) or cell.strip() != "")
            for cell in row
        ):
            count += 1

    return count


def compare_columns_between_files(
    file1: str,
    headers1: List[str],
    file2: str,
    headers2: List[str],
) -> Dict[str, Any]:
    """
    Compare columns from two Excel files.

    The function expects:
      - `headers1` and `headers2` are lists of equal length.
      - Columns are compared pairwise: headers1[i] <-> headers2[i].
      - Comparison is done row-by-row in order (first data row with first data row, etc.).
      - Values are treated as strings; missing cells are treated as empty string.

    Returns a dictionary with:
      - ok: bool (True if no mismatches)
      - rows_file1: int
      - rows_file2: int
      - mismatches: List[ { row_index: int, column_index: int, header1: str, header2: str,
                             value1: str, value2: str } ]

    Note: This is an order-sensitive comparison. If files must be aligned by a key (e.g. employee name),
    call read_excel and align rows before invoking this helper.
    """
    if len(headers1) != len(headers2):
        raise ValueError("headers1 and headers2 must have the same length")

    rows1 = read_excel(file1)
    rows2 = read_excel(file2)

    n1 = len(rows1)
    n2 = len(rows2)
    max_rows = max(n1, n2)

    mismatches = []

    def _as_str(v: Any) -> str:
        if v is None:
            return ""
        return str(v).strip()

    for row_idx in range(max_rows):
        row1 = rows1[row_idx] if row_idx < n1 else {}
        row2 = rows2[row_idx] if row_idx < n2 else {}
        for col_idx, (h1, h2) in enumerate(zip(headers1, headers2)):
            v1 = _as_str(row1.get(h1) if isinstance(row1, dict) else None)
            v2 = _as_str(row2.get(h2) if isinstance(row2, dict) else None)
            if v1 != v2:
                mismatches.append({
                    "row_index": row_idx + 1,            # 1-based data-row index (first data row is 1)
                    "column_index": col_idx,
                    "header1": h1,
                    "header2": h2,
                    "value1": v1,
                    "value2": v2,
                })

    return {
        "ok": len(mismatches) == 0,
        "rows_file1": n1,
        "rows_file2": n2,
        "mismatches": mismatches,
    }


def clean(v: Any) -> str:
    return "" if v is None else str(v).strip()


def normalize_header_name(h: Optional[str]) -> str:
    if h is None:
        return ""
    return "".join(ch for ch in h.lower() if ch.isalnum())


def find_key_header(headers: List[str], key_headers: Optional[List[str]] = None) -> Optional[str]:
    """
    Find a sensible key header for rows.

    Behavior:
      - If `key_headers` is provided, try its items in order and return the first matching header
        present in `headers`. Matching tries:
          1) whole-phrase regex (word-boundary) match against the header text,
          2) normalized-equality (remove non-alphameric, lowercase),
          3) token-presence (all tokens from candidate appear in normalized header).
      - If `key_headers` is not provided or none match, fall back to the legacy heuristics:
        prefer Officer Name, then Employee Name, then exact "Name", then Employer/Organization variants.

    This preserves backward compatibility: callers that pass only `headers` will get the same behavior.
    """
    if not headers:
        return None

    def _match_candidate(candidate: str) -> Optional[str]:
        if not candidate:
            return None
        cand_text = str(candidate).strip()
        cand_norm = normalize_header_name(cand_text)
        # 1) Whole-phrase match
        try:
            phrase_re = re.compile(r'\b' + re.escape(cand_text) + r'\b', re.IGNORECASE)
        except re.error:
            phrase_re = None
        if phrase_re:
            for h in headers:
                if h is None:
                    continue
                if phrase_re.search(str(h)):
                    return h
        # 2) Normalized equality
        for h in headers:
            if h is None:
                continue
            if normalize_header_name(h) == cand_norm:
                return h
        # 3) Token presence: all tokens appear in normalized header
        tokens = [t for t in re.split(r'\W+', cand_text.lower()) if t]
        if tokens:
            for h in headers:
                if h is None:
                    continue
                nh = normalize_header_name(h)
                if all(tok in nh for tok in tokens):
                    return h
        return None

    # If a prioritized list was supplied, try each candidate in order
    if key_headers:
        for cand in key_headers:
            found = _match_candidate(cand)
            if found:
                return found

    # Legacy heuristics (kept for backward compatibility)
    officer_re = re.compile(r'\bofficer\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            continue
        if officer_re.search(str(h)):
            return h

    for h in headers:
        if h is None:
            continue
        nh = normalize_header_name(h)
        if "officer" in nh and "name" in nh:
            return h

    employee_re = re.compile(r'\bemployee\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            continue
        if employee_re.search(str(h)):
            return h

    for h in headers:
        if h is None:
            continue
        if str(h).strip().lower() == "name":
            return h

    employer_re = re.compile(r'\bemployer\s+name\b', re.IGNORECASE)
    for h in headers:
        if h is None:
            continue
        if employer_re.search(str(h)):
            return h

    for h in headers:
        if h is None:
            continue
        nh = normalize_header_name(h)
        if "employer" in nh or "organization" in nh or ("org" in nh and "name" in nh) or ("organizationname" in nh):
            return h

    return None


def find_rank_header(headers: List[str]) -> Optional[str]:
    """
    Find a sensible rank-like header:
      - Rank Description variants
      - Employee Job Desc / Job Description variants
    """
    if not headers:
        return None

    patterns = [
        r'\brank\s+description\b',
        r'\brank\s+desc\b',
        r'\brankdescription\b',
        r'\bemployee\s+job\s+desc\b',
        r'\bemployee\s+job\s+description\b',
        r'\bjob\s+description\b',
        r'\bjob\s+desc\b',
    ]
    for pat in patterns:
        cre = re.compile(pat, re.IGNORECASE)
        for h in headers:
            if h is None:
                continue
            if cre.search(str(h)):
                return h

    for h in headers:
        if h is None:
            continue
        nh = normalize_header_name(h)
        if nh == "rankdescription" or ("rank" in nh and "description" in nh):
            return h
        if ("employeejob" in nh and ("desc" in nh or "description" in nh)) or (
            "job" in nh and ("desc" in nh or "description" in nh)
        ):
            return h

    return None


def build_map_by_key(rows: List[Dict[str, Any]], key_header: str) -> Dict[str, List[Dict[str, Any]]]:
    m: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        key = clean(r.get(key_header)) if isinstance(r, dict) else ""
        if key == "":
            key = f"__EMPTY_KEY__::{len(m)}"
        m.setdefault(key, []).append(r)
    return m


def is_numeric_str(x: str) -> bool:
    try:
        if x == "" or x is None:
            return False
        float(x)
        return True
    except Exception:
        return False


def is_zero_equivalent(x: str) -> bool:
    if x is None:
        return True
    s = str(x).strip()
    if s == "":
        return True
    try:
        return math.isclose(float(s), 0.0, rel_tol=1e-12, abs_tol=1e-9)
    except Exception:
        return False


def compare_rows_by_headers(
    r1: Dict[str, Any],
    r2: Dict[str, Any],
    headers: List[str],
    file1_name: str,
    file2_name: str,
    sheet: str,
    key: str,
    rank_header: Optional[str],
) -> List[str]:
    """
    Compare two row dicts across headers and return list of human-readable mismatch lines.
    Each line includes the sheet, key, Rank Description and the differing column + values.
    """
    diffs: List[str] = []

    rank_val = ""
    if rank_header and isinstance(r1, dict):
        rank_val = clean(r1.get(rank_header))
    if not rank_val and rank_header and isinstance(r2, dict):
        rank_val = clean(r2.get(rank_header))

    for h in headers:
        # ignore known noise columns by normalized name
        nh = normalize_header_name(h)
        if nh == "reporttime":
            continue
        if nh == "changetotal":
            continue
        if "ly" in nh and "total" in nh and "teamwork" in nh:
            continue
        if "ly" in nh and "total" in nh and any(tok in nh for tok in ("crit", "ast", "ovr", "unique")):
            continue
        if "change" in nh and "teamwork" in nh:
            continue
        if nh in ("change", "totalchange"):
            continue

        v1_raw = r1.get(h) if isinstance(r1, dict) else None
        v2_raw = r2.get(h) if isinstance(r2, dict) else None
        v1 = clean(v1_raw)
        v2 = clean(v2_raw)

        if is_zero_equivalent(v1) and is_zero_equivalent(v2):
            continue

        if is_numeric_str(v1) and is_numeric_str(v2):
            try:
                if math.isclose(float(v1), float(v2), rel_tol=1e-9, abs_tol=1e-6):
                    continue
            except Exception:
                pass

        if (is_zero_equivalent(v1) and is_numeric_str(v2) and math.isclose(float(v2), 0.0, rel_tol=1e-9, abs_tol=1e-6)) or (
            is_zero_equivalent(v2) and is_numeric_str(v1) and math.isclose(float(v1), 0.0, rel_tol=1e-9, abs_tol=1e-6)
        ):
            continue

        if v1 != v2:
            diffs.append(
                f"{sheet} | Key='{key}' | RankDescription='{rank_val}' | Column='{h}' | {file1_name}='{v1}' | {file2_name}='{v2}'"
            )
    return diffs